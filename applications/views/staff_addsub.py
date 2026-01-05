import json
import logging
import openpyxl
import qrcode
from django.shortcuts import render
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required
from ..firebase_service import FirebaseService
from collections import defaultdict
from io import BytesIO
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.views.decorators.cache import never_cache
from datetime import datetime, time

logger = logging.getLogger(__name__)


@login_required
@never_cache
def staff_addsub_view(request):
    try:
        firebase_service = FirebaseService()
        all_users = firebase_service.get_all_users()

        professors = []
        for user in all_users:
            first_name = user.get("firstName", "")
            mid_name = user.get("midName", "")
            last_name = user.get("lastName", "")
            middle_initial = f" {mid_name[0]}." if mid_name else ""
            full_name = f"{first_name}{middle_initial} {last_name}".strip()

            professors.append(
                {
                    "id": user.get("uid") or user.get("id"),
                    "full_name": full_name,
                    "email": user.get("email", ""),
                    "department": user.get("department", ""),
                    "role": user.get("role", ""),
                    "employID": user.get("employID", ""),
                }
            )

        professors.sort(key=lambda x: x["full_name"])
        return render(request, "staff/staff_addsub.html", {"professors": professors})

    except Exception as e:
        logger.error(f"Error loading semester view: {str(e)}")
        return render(request, "staff/staff_addsub.html", {"professors": []})


@require_POST
def add_class(request):
    try:
        data = json.loads(request.body)

        required_fields = [
            "subjectCode",
            "subjectName",
            "teacherUid",
            "day",
            "startTime",
            "endTime",
            "room",
        ]
        for field in required_fields:
            if not data.get(field):
                return JsonResponse({"error": f"{field} is required"}, status=400)

        firebase_service = FirebaseService()

        result = firebase_service.create_class(data)

        return JsonResponse(
            {
                "success": True,
                "message": "Class added successfully",
                "class": {
                    "id": result["class_id"],
                    "subjectCode": data["subjectCode"],
                    "subjectName": data["subjectName"],
                    "day": data["day"],
                    "startTime": data["startTime"],
                    "endTime": data["endTime"],
                    "room": data["room"],
                },
            }
        )

    except ValueError as ve:

        logger.warning(f"Schedule conflict: {str(ve)}")
        return JsonResponse({"success": False, "error": str(ve)}, status=409)
    except Exception as e:
        logger.error(f"Error adding class: {str(e)}")
        return JsonResponse({"success": False, "error": str(e)}, status=500)


def create_qr_class_code(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            text = data.get("text", "")

            if not text:
                return JsonResponse({"error": "No Text Provided"}, status=400)

            qr = qrcode.QRCode(
                version=1,
                error_correction=qrcode.constants.ERROR_CORRECT_L,
                box_size=10,
                border=4,
            )
            qr.add_data(text)
            qr.make(fit=True)

            img = qr.make_image(fill_color="black", back_color="white")

            img_io = BytesIO()
            img.save(img_io, "PNG")
            img_io.seek(0)

            response = HttpResponse(img_io, content_type="image/png")
            response["Content-Disposition"] = 'attachment; filename="qrcode.png"'
            return response

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse({"error": "Method not allowed"}, status=405)


@login_required
@require_POST
def import_class_excel(request):
    try:
        if "file" not in request.FILES:
            return JsonResponse({"error": "No file uploaded"}, status=400)

        excel_file = request.FILES["file"]
        wb = openpyxl.load_workbook(filename=BytesIO(excel_file.read()), data_only=True)
        sheet = wb.active

        teacherUid = request.POST.get("teacherUid")
        teacher_name = request.POST.get("teacher_name")
        department = request.POST.get("department", "all")

        class_list = []

        for row in sheet.iter_rows(min_row=2, values_only=True):

            if row and len(row) >= 2 and row[0] and row[1]:
                subjectCode = str(row[0]).strip()
                subjectName = str(row[1]).strip()
                day = str(row[2]).strip() if len(row) > 2 and row[2] else None
                startTime = str(row[3]).strip() if len(row) > 3 and row[3] else None
                endTime = str(row[4]).strip() if len(row) > 4 and row[4] else None
                room = str(row[5]).strip() if len(row) > 5 and row[5] else None

                class_list.append(
                    {
                        "subjectCode": subjectCode,
                        "subjectName": subjectName,
                        "teacherUid": teacherUid,
                        "teacher_name": teacher_name,
                        "department": department,
                        "day": day,
                        "startTime": startTime,
                        "endTime": endTime,
                        "room": room,
                    }
                )

        if not class_list:
            return JsonResponse(
                {"error": "No valid classes found in Excel file"}, status=400
            )

        firebase_service = FirebaseService()
        result = firebase_service.bulk_create_classes(class_list)

        message_parts = []
        if result["count"] > 0:
            message_parts.append(f"{result['count']} classes imported successfully")

        if result.get("skipped_count", 0) > 0:
            message_parts.append(
                f"{result['skipped_count']} classes skipped due to schedule conflicts"
            )

        message = ". ".join(message_parts)

        return JsonResponse(
            {
                "success": True,
                "message": message,
                "classes": result["classes"],
                "skipped": result.get("skipped", []),
                "imported_count": result["count"],
                "skipped_count": result.get("skipped_count", 0),
            }
        )

    except Exception as e:
        logger.error(f"Error importing Excel classes: {str(e)}")
        return JsonResponse({"success": False, "error": str(e)}, status=500)


@require_POST
def delete_course(request, class_id):
    try:
        firebase_service = FirebaseService()
        firebase_service.delete_class(class_id)
        return JsonResponse({"success": True, "message": "Class deleted successfully"})
    except Exception as e:
        logger.error(f"Error deleting class: {str(e)}")
        return JsonResponse({"success": False, "error": str(e)}, status=500)


@login_required
def get_classes_list(request):
    try:
        teacherUid = request.GET.get("teacherUid")
        department = request.GET.get("department", "all")

        logger.info(
            f"🔍 get_classes_list called with teacherUid={teacherUid}, department={department}"
        )

        firebase_service = FirebaseService()
        classes = firebase_service.get_classes(
            teacherUid=teacherUid, department=department
        )

        logger.info(f"📚 Returning {len(classes)} classes")
        return JsonResponse({"success": True, "classes": classes}, safe=False)
    except Exception as e:
        logger.error(f"Error fetching classes: {str(e)}")
        return JsonResponse({"success": False, "error": str(e)}, status=500)


@login_required
def get_user_profile(request):
    try:
        uid = request.GET.get("uid")
        if not uid:
            return JsonResponse({"success": False, "error": "UID required"}, status=400)

        firebase_service = FirebaseService()
        user = firebase_service.get_user_by_id(uid)

        if user:
            return JsonResponse({"success": True, "user": user})
        else:
            return JsonResponse(
                {"success": False, "error": "User not found"}, status=404
            )
    except Exception as e:
        logger.error(f"Error fetching user profile: {str(e)}")
        return JsonResponse({"success": False, "error": str(e)}, status=500)


@require_POST
def update_class(request, class_id):
    try:
        data = json.loads(request.body)

        if "startTime" in data and data["startTime"]:
            data["startTime"] = convert_to_ampm_format(data["startTime"])

        if "endTime" in data and data["endTime"]:
            data["endTime"] = convert_to_ampm_format(data["endTime"])

        firebase_service = FirebaseService()
        firebase_service.update_class(class_id, data)

        return JsonResponse(
            {
                "success": True,
                "message": "Class updated successfully",
                "class": {
                    "id": class_id,
                    "subjectCode": data.get("subjectCode"),
                    "subjectName": data.get("subjectName"),
                    "day": data.get("day"),
                    "startTime": data.get("startTime"),
                    "endTime": data.get("endTime"),
                    "room": data.get("room"),
                },
            }
        )
    except ValueError as ve:
        logger.warning(f"Schedule conflict during update: {str(ve)}")
        return JsonResponse({"success": False, "error": str(ve)}, status=409)
    except Exception as e:
        logger.error(f"Error updating class: {str(e)}")
        return JsonResponse({"success": False, "error": str(e)}, status=500)


def convert_to_ampm_format(time_str):
    """Convert time from various formats to AM/PM format"""
    if not time_str:
        return None

    time_str = str(time_str).strip()

    if "AM" in time_str.upper() or "PM" in time_str.upper():
        return time_str

    try:
        if ":" in time_str:
            parts = time_str.split(":")
            hours = int(parts[0])
            minutes = int(parts[1])

            period = "AM" if hours < 12 else "PM"
            display_hour = hours if hours <= 12 else hours - 12
            if display_hour == 0:
                display_hour = 12

            return f"{display_hour}:{minutes:02d} {period}"
    except (ValueError, IndexError):
        pass

    return time_str


def parse_excel_time(time_value):
    if time_value is None or str(time_value).strip() == "":
        return None

    if isinstance(time_value, time):
        hour = time_value.hour
        minute = time_value.minute
        period = "AM" if hour < 12 else "PM"
        display_hour = hour if hour <= 12 else hour - 12
        if display_hour == 0:
            display_hour = 12
        return f"{display_hour}:{minute:02d} {period}"

    if isinstance(time_value, datetime):
        hour = time_value.hour
        minute = time_value.minute
        period = "AM" if hour < 12 else "PM"
        display_hour = hour if hour <= 12 else hour - 12
        if display_hour == 0:
            display_hour = 12
        return f"{display_hour}:{minute:02d} {period}"

    if isinstance(time_value, (float, int)):

        total_seconds = int(time_value * 24 * 60 * 60)
        hour = (total_seconds // 3600) % 24
        minute = (total_seconds % 3600) // 60
        period = "AM" if hour < 12 else "PM"
        display_hour = hour if hour <= 12 else hour - 12
        if display_hour == 0:
            display_hour = 12
        return f"{display_hour}:{minute:02d} {period}"

    time_str = str(time_value).strip()

    if "AM" in time_str.upper() or "PM" in time_str.upper():
        try:

            period = "AM" if "AM" in time_str.upper() else "PM"

            time_part = time_str.upper().replace("AM", "").replace("PM", "").strip()

            parts = time_part.split(":")
            hour = int(parts[0])
            minute = int(parts[1]) if len(parts) > 1 else 0

            return f"{hour}:{minute:02d} {period}"
        except (ValueError, IndexError) as e:
            logger.warning(f"Error parsing time with AM/PM: {time_str}, error: {e}")
            return time_str

    try:
        if ":" in time_str:
            parts = time_str.split(":")
            hour = int(parts[0])
            minute = int(parts[1])

            period = "AM" if hour < 12 else "PM"
            display_hour = hour if hour <= 12 else hour - 12
            if display_hour == 0:
                display_hour = 12
            return f"{display_hour}:{minute:02d} {period}"
    except (ValueError, IndexError):
        pass

    logger.warning(
        f"Could not parse time value: {time_value} (type: {type(time_value)})"
    )
    return str(time_value)


@require_http_methods(["POST"])
def import_all_teachers_excel(request):
    try:

        if "file" not in request.FILES:
            return JsonResponse({"error": "No file uploaded"}, status=400)

        excel_file = request.FILES["file"]

        wb = openpyxl.load_workbook(filename=BytesIO(excel_file.read()), data_only=True)
        sheet = wb.active

        teachers_classes = defaultdict(list)
        errors = []

        current_employee_id = None
        in_data_section = False
        row_number = 0

        for row in sheet.iter_rows(values_only=True):
            row_number += 1

            if not row or not any(row):
                in_data_section = False
                continue

            row_data = list(row)

            logger.info(
                f"Row {row_number}: {[str(cell)[:30] if cell else 'None' for cell in row_data[:10]]}"
            )

            col_b = str(row_data[1]).strip().lower() if len(row_data) > 1 else ""

            if "employee" in col_b and "id" in col_b:

                id_col = 2

                employee_id_value = row_data[id_col] if len(row_data) > id_col else None

                if not employee_id_value:
                    logger.warning(
                        f"Row {row_number}: Employee ID value is missing, skipping"
                    )
                    continue

                current_employee_id = str(employee_id_value).strip()
                if "." in current_employee_id:
                    current_employee_id = current_employee_id.split(".")[0]

                in_data_section = False

                logger.info(
                    f"Row {row_number}: ✓ Found Employee ID: {current_employee_id}"
                )

            elif "subject" in col_b and "code" in col_b:
                in_data_section = True
                logger.info(
                    f"Row {row_number}: ✓ Starting data section for Employee ID: {current_employee_id}"
                )
                continue

            elif in_data_section and current_employee_id:

                subject_code = str(row_data[1]).strip() if len(row_data) > 1 else ""
                subject_name = str(row_data[2]).strip() if len(row_data) > 2 else ""

                if not subject_code or not subject_name:
                    logger.info(
                        f"Row {row_number}: Empty or incomplete - ending data section"
                    )
                    in_data_section = False
                    continue

                if "subject" in subject_code.lower() or "code" in subject_code.lower():
                    logger.info(f"Row {row_number}: Duplicate header, skipping")
                    continue

                try:
                    day = str(row_data[3]).strip() if len(row_data) > 3 else ""
                    section = str(row_data[4]).strip() if len(row_data) > 4 else ""

                    start_time_raw = row_data[5] if len(row_data) > 5 else None
                    end_time_raw = row_data[6] if len(row_data) > 6 else None

                    start_time = parse_excel_time(start_time_raw)
                    end_time = parse_excel_time(end_time_raw)

                    room = str(row_data[7]).strip() if len(row_data) > 7 else ""

                    if not start_time or not end_time:
                        error_msg = f"Row {row_number}: Missing time for {subject_code} - {subject_name}"
                        errors.append(error_msg)
                        logger.warning(error_msg)
                        continue

                    class_data = {
                        "subjectCode": subject_code,
                        "subjectName": subject_name,
                        "day": day,
                        "section": section,
                        "startTime": start_time,
                        "endTime": end_time,
                        "room": room,
                    }

                    teachers_classes[current_employee_id].append(class_data)
                    logger.info(
                        f"Row {row_number}: ✓ Added {subject_code} ({subject_name}) for {current_employee_id} ({start_time} - {end_time})"
                    )

                except Exception as e:
                    error_msg = f"Row {row_number}: Error parsing class - {str(e)}"
                    errors.append(error_msg)
                    logger.error(f"{error_msg}, Data: {row_data[:10]}")
                    continue

        logger.info(f"Parsed {len(teachers_classes)} teachers with classes")
        for emp_id, classes in teachers_classes.items():
            logger.info(f"  {emp_id}: {len(classes)} classes")

        if not teachers_classes:
            return JsonResponse(
                {
                    "error": "No valid classes found in Excel file",
                    "errors": errors,
                    "hint": "Make sure your Excel file has the correct format with Employee ID and class data.",
                },
                status=400,
            )

        firebase_service = FirebaseService()

        teacher_results = []
        total_imported = 0
        total_skipped = 0
        skipped_details = []

        for employee_id, class_list in teachers_classes.items():
            try:

                teacher_info = firebase_service.get_teacher_by_employee_id(employee_id)

                if not teacher_info:
                    errors.append(
                        f"Employee ID: {employee_id} not found in system. "
                        f"Skipped {len(class_list)} classes."
                    )
                    logger.warning(
                        f"Teacher with Employee ID {employee_id} not found in Firebase"
                    )

                    total_skipped += len(class_list)

                    for class_data in class_list:
                        skipped_details.append(
                            {
                                "employee_id": employee_id,
                                "teacher_name": f"ID-{employee_id}",
                                "subjectCode": class_data.get("subjectCode", "N/A"),
                                "reason": "Teacher not found in system",
                            }
                        )

                    continue

                teacher_uid = teacher_info.get("uid")
                display_name = teacher_info.get("name", f"ID-{employee_id}")

                for class_data in class_list:
                    class_data["teacherUid"] = teacher_uid
                    class_data["teacher_name"] = display_name
                    class_data["department"] = teacher_info.get("department", "all")
                    class_data["employee_id"] = employee_id

                result = firebase_service.bulk_create_classes(class_list)

                teacher_results.append(
                    {
                        "employee_id": employee_id,
                        "teacher_name": display_name,
                        "imported_count": result["count"],
                        "skipped_count": result.get("skipped_count", 0),
                    }
                )

                total_imported += result["count"]
                total_skipped += result.get("skipped_count", 0)

                if result.get("skipped"):
                    for skipped in result["skipped"]:
                        skipped_details.append(
                            {
                                "employee_id": employee_id,
                                "teacher_name": display_name,
                                "subjectCode": skipped["subjectCode"],
                                "reason": skipped["reason"],
                            }
                        )

                logger.info(
                    f"Processed {employee_id}: {result['count']} imported, {result.get('skipped_count', 0)} skipped"
                )

            except Exception as e:
                logger.error(
                    f"Error processing teacher {employee_id}: {str(e)}", exc_info=True
                )
                errors.append(f"Error processing Employee ID ({employee_id}): {str(e)}")
                continue

        response_data = {
            "success": True,
            "teachers_processed": len(teacher_results),
            "total_imported": total_imported,
            "total_skipped": total_skipped,
            "teacher_results": teacher_results,
            "skipped_details": skipped_details,
            "errors": errors,
            "has_errors": len(errors) > 0 or total_skipped > 0,
        }

        logger.info(
            f"Import complete: {total_imported} imported, {total_skipped} skipped, {len(errors)} errors"
        )
        return JsonResponse(response_data)

    except Exception as e:
        logger.error(f"Error importing all teachers' Excel: {str(e)}", exc_info=True)
        return JsonResponse(
            {"success": False, "error": f"Failed to process file: {str(e)}"}, status=500
        )
